import openpyxl
from openpyxl.styles import Font, PatternFill
# 1. Crea un nuevo libro de excel y seleccionar la hoja activa
wb = openpyxl.Workbook()
hoja = wb.active
hoja.title =  "Reporte de venta"
# 2. Agragar los encabezados de la tabla
hoja['A1'] = "Producto"
hoja['B1'] = "Cantidad"
hoja['C1'] = "Precio Unitario"
hoja['D1'] = "Total"

# 3. Datos simulados para la automatizar
datos_ventas=[
    ["Laptop Dell", 2, 800],
    ["Monitor Asus", 5, 250],
    ["Teclado Mecanico", 10, 70],
    ["Mouse Gamer", 15, 45]
]

# 4. Llenar la tabla e insertar formulas de Excel automaticamente
for fila, producto in enumerate ( datos_ventas, start=2):
    hoja[f'A{fila}'] = producto[0]
    hoja[f'B{fila}'] = producto[1]
    hoja[f'C{fila}'] = producto[2]
    # Automatizamos la fórmula de multiplicación de Excel: = Cantidad * Precio
    hoja[f'D{fila}'] = f"=B{fila}*C{fila}"

# 5. Agregar la fila de " Gran Total" al final
ultima_fila = len(datos_ventas) + 2
hoja[f'C{ultima_fila}'] = "Gran Total:"
hoja[f'D{ultima_fila}'] = f'=SUM("D2:D{ultima_fila-1}")'

# 6. Darle un toque profesional(Estilos)
fuente_titulo = Font (name= "Arial", size=12, bold = True,color= "FFFFFF" )
fondo_azul = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type="solid")

# Aplicar estilo en el encabezado
for columna in ['A1', 'B1', 'C1', 'D1']:
    hoja[columna].font = fuente_titulo
    hoja[columna].fill=fondo_azul

# 7. Guardar el archivo en tu computadora
    nombre_archivo= "Reporte_Automatizado.xlsx"
    wb.save(nombre_archivo)
    print(f"¡Éxito! El archivo '{nombre_archivo}' ha sido creado y automatizado.")
